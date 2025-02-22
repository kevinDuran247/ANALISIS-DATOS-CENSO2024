import openpyxl

def get_data_poblacion(file_path="../db/poblacion.xlsx"):

    excel_dataframe = openpyxl.load_workbook(file_path)
    dataframe = excel_dataframe.active

    data = []
    
    distribucion_total = 0.0
    veces = 0
    

    suma_columna_3 = sum(
        row[2].value for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row) 
        if isinstance(row[2].value, (int, float)) 
    )

    for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row, max_col=dataframe.max_column):
        _row = [cell.value for cell in row] 
        
        if(veces == 0 ):
            fila_4 = "Distribucion Total"
            veces += 1
        
        elif isinstance(row[2].value, (int, float)) and suma_columna_3 != 0:
            fila_4 = row[2].value / suma_columna_3
            distribucion_total +=  row[2].value / suma_columna_3
        else:
            fila_4 = None  

        
        _row.append(fila_4) 
        del _row[4]
        data.append(_row)
    
    suma_columna_2 = sum(row[1] for row in data if isinstance(row[1], (int, float)))
    suma_columna_4 = sum(row[3] for row in data if isinstance(row[1], (int, float)))
   

    fila_personalizada = ["Nacional", suma_columna_2,suma_columna_3,suma_columna_4, distribucion_total] + [None] * (dataframe.max_column - 5)
    
    data.append(fila_personalizada)

    return data
