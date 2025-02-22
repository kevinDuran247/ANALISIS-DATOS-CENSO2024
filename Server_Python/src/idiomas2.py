import openpyxl

def get_data_idiomas2(file_path="../db/idiomas2.xlsx"):
    excel_dataframe = openpyxl.load_workbook(file_path)
    dataframe = excel_dataframe.active

    data = []
    
   
    for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row, max_col=dataframe.max_column):
        _row = [cell.value for cell in row] 
        
       
        data.append(_row)
    
  

    return data
