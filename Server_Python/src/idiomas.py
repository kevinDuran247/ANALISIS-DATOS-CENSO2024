import openpyxl

def get_data_idiomas(file_path="../db/idiomas.xlsx"):
    # Cargar el archivo de Excel
    excel_dataframe = openpyxl.load_workbook(file_path)
    dataframe = excel_dataframe.active

    data = []
    

    for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row, max_col=dataframe.max_column):
        _row = [cell.value for cell in row] 
        
     
        data.append(_row)
    
    suma_columna_2 = sum(row[1] for row in data if isinstance(row[1], (int, float)))
    suma_columna_4 = sum(row[2] for row in data if isinstance(row[1], (int, float)))
    suma_columna_5 = sum(row[3] for row in data if isinstance(row[1], (int, float)))
   

    fila_personalizada = ["Nacional", suma_columna_2,suma_columna_4,suma_columna_5] + [None] * (dataframe.max_column - 4)
    
    data.append(fila_personalizada)

    return data
