import openpyxl

def get_data_discapacidades(file_path="../db/discapacidades.xlsx"):

    excel_dataframe = openpyxl.load_workbook(file_path)
    dataframe = excel_dataframe.active

    data = []
    
    distribucion_total = 0.0
    veces = 0
    

    suma_columna_8 = sum(
        row[8].value for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row) 
        if isinstance(row[8].value, (int, float)) 
    )

    for row in dataframe.iter_rows(min_row=2, max_row=dataframe.max_row, max_col=dataframe.max_column):
        _row = [cell.value for cell in row] 
        
        if(veces == 0 ):
            fila_8 = "Distribucion Total"
            veces += 1
        
        elif isinstance(row[8].value, (int, float)) and suma_columna_8 != 0:
            fila_8 = row[8].value / suma_columna_8
            distribucion_total +=  row[8].value / suma_columna_8
        else:
            fila_8 = None  

        
        _row.append(fila_8) 
        del _row[9]
        data.append(_row)
    
    suma_columna_1 = sum(row[1] for row in data if isinstance(row[1], (int, float)))
    suma_columna_2 = sum(row[2] for row in data if isinstance(row[1], (int, float)))
    suma_columna_3 = sum(row[3] for row in data if isinstance(row[1], (int, float)))
    suma_columna_4 = sum(row[4] for row in data if isinstance(row[1], (int, float)))
    suma_columna_5 = sum(row[5] for row in data if isinstance(row[1], (int, float)))
    suma_columna_6 = sum(row[6] for row in data if isinstance(row[1], (int, float)))
    suma_columna_7 = sum(row[7] for row in data if isinstance(row[1], (int, float)))
   

    fila_personalizada = ["Nacional",suma_columna_1, suma_columna_2, suma_columna_3,suma_columna_4,suma_columna_5,suma_columna_6,suma_columna_7,suma_columna_8,distribucion_total] + [None] * (dataframe.max_column - 10)
    
    data.append(fila_personalizada)

    return data
